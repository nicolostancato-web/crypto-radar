"""
jobs/export_excel.py — la VETRINA.

Scrive in un .xlsx i top N asset per score, ordinati dal più alto.
È il file che apri tu. NON contiene "segnali da comprare": contiene candidati da
GUARDARE, con il PERCHÉ del voto e il prezzo al momento, così puoi fare paper
trading onesto (segni entrata ipotetica, controlli tra qualche giorno).

Colonne pensate per smascherare l'autoinganno:
  - price_at_score: prezzo quando lo score è scattato (il tuo punto di entrata ipotetico)
  - breakdown: da dove viene il voto (niente numeri magici)
  - age_hours / liquidity: per ricordarti che gli illiquidi non si tradano davvero
"""
import sys, os, time, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import SCORING, OUTCOMES, EXPORT_PATH
from db import get_conn, init_db


def _score_10(raw):
    """Mappa lo score grezzo su scala 0-10 (la verità resta il grezzo nel breakdown)."""
    ref = SCORING.get("score_reference", 12.0)
    return round(min(10.0, max(0.0, raw / ref * 10.0)), 1)


def _append_validation_sheet(path):
    """Aggiunge la scheda 'Validazione': il VERDETTO coi dati (valore atteso netto)."""
    from openpyxl import load_workbook
    with get_conn() as c:
        rows = c.execute("SELECT * FROM outcomes ORDER BY entered_at DESC").fetchall()

    wb = load_workbook(path)
    ws = wb.create_sheet("Validazione")
    bold = Font(bold=True)
    title = Font(bold=True, color="FFFFFF", name="Arial")
    fill = PatternFill("solid", start_color="1F4E78")

    ws["A1"] = "VALIDAZIONE — gli score alti si muovono DAVVERO? (rendimento NETTO di slippage+fee)"
    ws["A1"].font = bold
    ws["A3"] = "Orizzonte"; ws["B3"] = "N"; ws["C3"] = "Valore atteso NETTO"; ws["D3"] = "Win rate"
    for col in "ABCD":
        ws[f"{col}3"].font = title; ws[f"{col}3"].fill = fill
    r = 4
    for h in OUTCOMES["horizons"]:
        nets = [row[f"ret_{h}h_net"] for row in rows if row[f"ret_{h}h_net"] is not None]
        ws[f"A{r}"] = f"{h}h"
        if nets:
            avg = sum(nets) / len(nets)
            win = sum(1 for x in nets if x > 0) / len(nets)
            ws[f"B{r}"] = len(nets)
            ws[f"C{r}"] = f"{avg:+.2%}"
            ws[f"D{r}"] = f"{win:.0%}"
        else:
            ws[f"B{r}"] = 0; ws[f"C{r}"] = "—ancora maturo"; ws[f"D{r}"] = "—"
        r += 1

    # elenco entrate (le piu' recenti)
    r += 1
    headers = ["Ticker", "Score_entry", "Prezzo_entry", "Ret_72h_netto", "Stato", "Entrata"]
    for i, hd in enumerate(headers):
        cell = ws.cell(row=r, column=1 + i, value=hd)
        cell.font = title; cell.fill = fill
    r += 1
    import datetime
    for row in rows[:50]:
        ret72 = row["ret_72h_net"]
        ws.cell(row=r, column=1, value=row["ticker"])
        ws.cell(row=r, column=2, value=round(row["score_at_entry"] or 0, 2))
        ws.cell(row=r, column=3, value=row["price_at_entry"])
        ws.cell(row=r, column=4, value=(f"{ret72:+.2%}" if ret72 is not None else "—"))
        ws.cell(row=r, column=5, value=row["status"])
        ws.cell(row=r, column=6,
                value=datetime.datetime.fromtimestamp(row["entered_at"]).strftime("%m-%d %H:%M"))
        r += 1

    for col, w in {"A": 16, "B": 12, "C": 20, "D": 12, "E": 10, "F": 14}.items():
        ws.column_dimensions[col].width = w
    wb.save(path)


def export():
    init_db()
    with get_conn() as c:
        rows = c.execute(
            """SELECT a.ticker, a.name, a.chain, a.contract_address,
                      s.current_score, s.price_at_score, s.breakdown, s.updated_at,
                      a.discovered_at, a.discovery_source
               FROM scores s JOIN assets a ON a.id = s.asset_id
               WHERE a.status='active' AND s.current_score >= ?
               ORDER BY s.current_score DESC
               LIMIT ?""",
            (SCORING["min_score_for_export"], SCORING["top_n_export"]),
        ).fetchall()

    now = time.time()
    data = []
    for r in rows:
        bd = json.loads(r["breakdown"] or "{}")
        bd_str = ", ".join(f"{k}:{v}" for k, v in sorted(bd.items(), key=lambda x: -x[1]))
        data.append({
            "Score_su_10": _score_10(r["current_score"]),
            "Ticker": r["ticker"],
            "Nome": r["name"],
            "Score_grezzo": round(r["current_score"], 2),
            "Prezzo_al_voto_USD": r["price_at_score"],
            "Perche_score": bd_str,
            "Chain": r["chain"],
            "Eta_ore": round((now - r["discovered_at"]) / 3600, 1),
            "Aggiornato_min_fa": round((now - r["updated_at"]) / 60, 1),
            "Fonte": r["discovery_source"],
            "Contract": r["contract_address"],
            # colonne VUOTE che compili tu a mano durante il paper trading:
            "Entrata_ipotetica": "",
            "Prezzo_dopo_3g": "",
            "Esito": "",
        })

    # FIX "Excel stale": se non c'è nulla sopra soglia, scriviamo COMUNQUE un file con
    # un avviso datato, così non resta a video un pick vecchio scaduto.
    if not data:
        import datetime
        stamp = datetime.datetime.fromtimestamp(now).strftime("%Y-%m-%d %H:%M")
        df = pd.DataFrame([{
            "Score_su_10": "", "Ticker": "— nessun candidato sopra soglia —",
            "Nome": f"ultimo controllo: {stamp}", "Score_grezzo": "",
            "Perche_score": "il sistema sta osservando; nessuna confluenza forte ora",
        }])
        df.to_excel(EXPORT_PATH, index=False, sheet_name="Top Scores")
        _append_validation_sheet(EXPORT_PATH)
        print(f"[export] nessun asset sopra soglia: scritto avviso datato ({stamp}).")
        return EXPORT_PATH

    df = pd.DataFrame(data)
    df.to_excel(EXPORT_PATH, index=False, sheet_name="Top Scores")

    # formattazione leggera (openpyxl)
    from openpyxl import load_workbook
    wb = load_workbook(EXPORT_PATH)
    ws = wb["Top Scores"]
    header_fill = PatternFill("solid", start_color="1F4E78")
    for col_idx, _ in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # larghezza colonna ~ contenuto
        col = get_column_letter(col_idx)
        maxlen = max([len(str(df.columns[col_idx-1]))] +
                     [len(str(v)) for v in df.iloc[:, col_idx-1].tolist()])
        ws.column_dimensions[col].width = min(maxlen + 2, 45)
    ws.freeze_panes = "A2"
    wb.save(EXPORT_PATH)
    _append_validation_sheet(EXPORT_PATH)
    print(f"[export] {len(df)} righe scritte in {EXPORT_PATH}")
    return EXPORT_PATH


if __name__ == "__main__":
    export()
